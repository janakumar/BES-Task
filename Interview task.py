import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class FileProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("CSV to Excel Processor")
        
        self.data1_path = ""
        self.desc_path = ""
        self.boq_path = ""
        
        tk.Label(root, text="Select Data1 CSV:").grid(row=0, column=0)
        self.data1_btn = tk.Button(root, text="Browse", command=self.load_data1)
        self.data1_btn.grid(row=0, column=1)
        
        tk.Label(root, text="Select Description CSV:").grid(row=1, column=0)
        self.desc_btn = tk.Button(root, text="Browse", command=self.load_desc)
        self.desc_btn.grid(row=1, column=1)
        
        tk.Label(root, text="Select BOQ Excel File:").grid(row=2, column=0)
        self.boq_btn = tk.Button(root, text="Browse", command=self.load_boq)
        self.boq_btn.grid(row=2, column=1)
        
        self.process_btn = tk.Button(root, text="Process and Export", command=self.process_files)
        self.process_btn.grid(row=3, column=0, columnspan=2)
    
    def load_data1(self):
        self.data1_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    
    def load_desc(self):
        self.desc_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    
    def load_boq(self):
        self.boq_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    
    def update_excel_with_format(self, boq_path, updated_df):
        # Load existing workbook
        wb = load_workbook(boq_path)
        ws = wb.active  
        merged_ranges = ws.merged_cells.ranges  
        rows, cols = updated_df.shape
        for i in range(rows):
            for j in range(cols):
                cell = ws.cell(row=i+2, column=j+1) 
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        if cell.coordinate == merged_range.start_cell.coordinate:
                            cell.value = updated_df.iloc[i, j]
                        break  
                else:
                    cell.value = updated_df.iloc[i, j]
        wb.save(boq_path)
    
    def process_files(self):
        if not self.data1_path or not self.desc_path or not self.boq_path:
            messagebox.showerror("Error", "Please select all files")
            return
        
        try:
            # Load data
            boq_df = pd.read_excel(self.boq_path, sheet_name=0) 
            data1_df = pd.read_csv(self.data1_path)
            desc_df = pd.read_csv(self.desc_path)

            # Renaming columns in Data1 to match BOQ column names
            data1_df.rename(columns={"RM (m)": "RMT"}, inplace=True)
            columns_to_merge = ["Facade Type", "Sub Type", "Area (Sqm)", "RMT", "Glass (Sqm)", "System (EOI) (Sqm)", "Steel (Tonnes)"]
            data1_df = data1_df[columns_to_merge]
            boq_df.update(data1_df, overwrite=False)

            self.update_excel_with_format(self.boq_path, boq_df)

            # Second half of the processing
            boq_df = pd.read_excel(self.boq_path, sheet_name=0)
            desc_df.rename(columns={"Description": "Brief Description", "Cost": "Rate\n(INR)"}, inplace=True)
            columns_to_merge = ["Facade Type", "Sub Type", "Facade System", "Brief Description", "Rate\n(INR)"]
            desc_df = desc_df[columns_to_merge]
            boq_df.update(desc_df, overwrite=False)
            boq_df.fillna("#REF!",inplace=True)

            self.update_excel_with_format(self.boq_path, boq_df)

            messagebox.showinfo("Success", "BOQ file updated successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = FileProcessor(root)
    root.mainloop()